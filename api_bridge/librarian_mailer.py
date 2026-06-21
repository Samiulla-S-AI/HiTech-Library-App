"""
HiTECH Library — Librarian Email Manager
A CustomTkinter desktop app for the librarian to review and send
Reminder & Overdue emails to students. Windows 7+ compatible.

Features:
  - Reminder tab: students whose books are due tomorrow
  - Today's Due tab: students whose books are due today (checks if reminded yesterday)
  - Overdue tab: students with overdue books (sortable by days, dept, title, year)
  - Settings: reg-no prefix sequences to filter active students
  - Multi-select + Send All
  - Local SQLite DB for sent-mail history & settings
  - History viewer with delete
"""

import os
import sys
import json
import sqlite3
import threading
import smtplib
import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

import csv
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
import customtkinter as ctk  # type: ignore
from tkinter import messagebox, END, filedialog

# Thread lock for SQLite database to prevent concurrent write errors
local_db_lock = threading.Lock()

# ─── paths ───────────────────────────────────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOCAL_DB_PATH = os.path.join(SCRIPT_DIR, "librarian_mailer.db")

# In Windows, if the script is placed in a restricted folder without write 
# permissions, SQLite will fail to open the db and throw an error. 
# We test if we can write to the DB. If not, we fallback to the user's home directory.
try:
    with open(LOCAL_DB_PATH, 'a') as _f:
        pass
except (IOError, OSError, PermissionError):
    _fallback_dir = os.path.join(os.path.expanduser("~"), "Library_APP_Data")
    if not os.path.exists(_fallback_dir):
        try:
            os.makedirs(_fallback_dir)
        except Exception:
            pass
    LOCAL_DB_PATH = os.path.join(_fallback_dir, "librarian_mailer.db")

# ─── load .env manually (no dotenv dependency) ──────────────────────────
def load_env(path):
    if not os.path.exists(path):
        return {}
    env = {}
    with open(path, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                env[k.strip()] = v.strip()
    return env

ENV = load_env(os.path.join(SCRIPT_DIR, ".env"))

DB_SERVER = ENV.get("DB_SERVER", "172.16.0.114")
DB_NAME = ENV.get("DB_NAME", "lips5_sql")
DB_USER = ENV.get("DB_USER", "LibraryUserReader1")
DB_PASSWORD = ENV.get("DB_PASSWORD", "hitlibrary@123")
GMAIL_USER = ENV.get("GMAIL_USER", "gigwork.samiulla@gmail.com")
GMAIL_APP_PASSWORD = ENV.get("GMAIL_APP_PASSWORD", "oyvszyiwoxtxxaiy")



# ══════════════════════════════════════════════════════════════════════════
#  LOCAL SQLITE — settings, history, sent status
# ══════════════════════════════════════════════════════════════════════════
def init_local_db():
    with local_db_lock:
        conn = sqlite3.connect(LOCAL_DB_PATH)
        c = conn.cursor()
        c.execute("""
            CREATE TABLE IF NOT EXISTS settings (
                key TEXT PRIMARY KEY,
                value TEXT
            )
        """)
        c.execute("""
            CREATE TABLE IF NOT EXISTS sent_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                roll_no TEXT,
                name TEXT,
                email TEXT,
                mail_type TEXT,
                subject TEXT,
                sent_at TEXT,
                details TEXT
            )
        """)
        # Add indexes for performance
        c.execute("CREATE INDEX IF NOT EXISTS idx_sent_roll ON sent_history(roll_no)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_sent_type ON sent_history(mail_type)")
        c.execute("CREATE INDEX IF NOT EXISTS idx_sent_at ON sent_history(sent_at)")
        conn.commit()
        conn.close()


def get_setting(key, default=""):
    with local_db_lock:
        conn = sqlite3.connect(LOCAL_DB_PATH)
        c = conn.cursor()
        c.execute("SELECT value FROM settings WHERE key=?", (key,))
        row = c.fetchone()
        conn.close()
    return row[0] if row else default


def set_setting(key, value):
    with local_db_lock:
        conn = sqlite3.connect(LOCAL_DB_PATH)
        c = conn.cursor()
        c.execute("INSERT OR REPLACE INTO settings(key,value) VALUES(?,?)", (key, value))
        conn.commit()
        conn.close()


def add_sent_record(roll_no, name, email, mail_type, subject, details=""):
    # Use a longer timeout to prevent 'database is locked' or 'readonly' errors during concurrent writes
    with local_db_lock:
        conn = sqlite3.connect(LOCAL_DB_PATH, timeout=20)
        try:
            c = conn.cursor()
            c.execute(
                "INSERT INTO sent_history(roll_no,name,email,mail_type,subject,sent_at,details) VALUES(?,?,?,?,?,?,?)",
                (roll_no, name, email, mail_type, subject,
                 datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"), details),
            )
            conn.commit()
        finally:
            conn.close()


def was_sent_today(roll_no, mail_type):
    today = datetime.date.today()
    return was_sent_on(roll_no, mail_type, today)


def was_sent_on(roll_no, mail_type, date_obj):
    """Check if an email of specific type was sent to this roll_no on a specific date."""
    with local_db_lock:
        conn = sqlite3.connect(LOCAL_DB_PATH)
        c = conn.cursor()
        date_str = date_obj.isoformat()
        # Handle multiple mail types (e.g. 'reminder' and 'today_reminder' are both reminders)
        if mail_type == "reminder":
            c.execute(
                "SELECT COUNT(*) FROM sent_history WHERE roll_no=? AND mail_type IN ('reminder','today_reminder') AND sent_at LIKE ?",
                (roll_no, f"{date_str}%"),
            )
        else:
            c.execute(
                "SELECT COUNT(*) FROM sent_history WHERE roll_no=? AND mail_type=? AND sent_at LIKE ?",
                (roll_no, mail_type, f"{date_str}%"),
            )
        cnt = c.fetchone()[0]
        conn.close()
    return cnt > 0


def get_all_sent_history_counts():
    """Returns a dict mapping roll_no -> list of sent_at datetime strings."""
    with local_db_lock:
        conn = sqlite3.connect(LOCAL_DB_PATH)
        c = conn.cursor()
        c.execute("SELECT roll_no, sent_at FROM sent_history ORDER BY sent_at DESC")
        rows = c.fetchall()
        conn.close()
    
    h_map = {}
    for r, dt in rows:
        if r not in h_map:
            h_map[r] = []
        h_map[r].append(dt)
    return h_map


def get_history(limit=200):
    with local_db_lock:
        conn = sqlite3.connect(LOCAL_DB_PATH)
        c = conn.cursor()
        c.execute("SELECT id,roll_no,name,email,mail_type,subject,sent_at,details FROM sent_history ORDER BY id DESC LIMIT ?", (limit,))
        rows = c.fetchall()
        conn.close()
    return rows


def delete_history_ids(ids):
    if not ids:
        return
    conn = sqlite3.connect(LOCAL_DB_PATH)
    c = conn.cursor()
    c.execute(f"DELETE FROM sent_history WHERE id IN ({','.join('?' * len(ids))})", ids)
    conn.commit()
    conn.close()


# ══════════════════════════════════════════════════════════════════════════
def mssql_connect():
    import pymssql  # type: ignore
    # Clean up the server string (handle the comma from .env)
    clean_server = DB_SERVER.split(',')[0].strip()
    
    # Try different TDS versions that Windows Server 2008 R2 likes
    # We try 7.0 first as it's the most compatible for old SQL Servers
    for tds in ["7.0", "7.1", "7.2"]:
        try:
            print(f"📡 Attempting connection (IP: {clean_server} | TDS: {tds})...")
            return pymssql.connect(
                server=clean_server, 
                user=DB_USER, 
                password=DB_PASSWORD,
                database=DB_NAME, 
                port=1433, 
                login_timeout=10, 
                charset="UTF-8",
                tds_version=tds
            )
        except Exception as e:
            # If it's a connection failure, try the next TDS version
            if "Adaptive Server connection failed" in str(e) and tds != "7.2":
                print(f"⚠️ TDS {tds} failed, trying next...")
                continue
            
            # If all failed, print everything
            print(f"\n❌ [CONNECTION FAILED]")
            print(f"Final Attempt IP: {clean_server}")
            print(f"Error: {str(e)}")
            print("-" * 30)
            print("SUGGESTION: If this fails, go to Computer A and")
            print("change 'TCP Dynamic Ports' in IPAll from 49169 to BLANK.")
            raise e



def fetch_departments():
    """Fetch all departments from the database."""
    conn = mssql_connect()
    if conn is None: return []
    cur = conn.cursor(as_dict=True)
    cur.execute("SELECT dept_no, dept_name FROM [dbo].[department] ORDER BY dept_name")
    rows = cur.fetchall()
    conn.close()
    return rows


def fetch_reminder_students(prefixes):
    """Students whose book(s) are due TOMORROW, filtered by prefixes."""
    conn = mssql_connect()
    if conn is None: return []
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT
            bc.id_no        AS RollNumber,
            p.name          AS Name,
            ISNULL(p.e_mail,'') AS Email,
            ISNULL(d.dept_name,'') AS Department,
            b.title         AS BookTitle,
            ISNULL(b.author, '') AS Author,
            bc.acc_no       AS AccNo,
            bc.t_date       AS IssueDate,
            bc.due_date     AS DueDate,
            ISNULL(c.cat_name,'') AS Category
        FROM [dbo].[book_circle] bc
        LEFT JOIN [dbo].[book]       b ON bc.acc_no = CAST(b.acc_no AS VARCHAR)
        LEFT JOIN [dbo].[personal]   p ON bc.id_no  = p.id_no
        LEFT JOIN [dbo].[department] d ON p.dept_no = d.dept_no
        LEFT JOIN [dbo].[catagory]   c ON p.cat_no  = c.cat_no
        WHERE CAST(bc.due_date AS DATE) = CAST(DATEADD(day,1,GETDATE()) AS DATE)
          AND p.e_mail IS NOT NULL AND p.e_mail != ''
        ORDER BY p.name, bc.due_date
    """)
    rows = cur.fetchall()
    conn.close()
    return _filter_prefixes(rows, prefixes)


def fetch_today_students(prefixes):
    """Students whose book(s) are due TODAY, filtered by prefixes."""
    conn = mssql_connect()
    if conn is None: return []
    cur = conn.cursor(as_dict=True)
    cur.execute("""
        SELECT
            bc.id_no        AS RollNumber,
            p.name          AS Name,
            ISNULL(p.e_mail,'') AS Email,
            ISNULL(d.dept_name,'') AS Department,
            b.title         AS BookTitle,
            ISNULL(b.author, '') AS Author,
            bc.acc_no       AS AccNo,
            bc.t_date       AS IssueDate,
            bc.due_date     AS DueDate,
            ISNULL(c.cat_name,'') AS Category
        FROM [dbo].[book_circle] bc
        LEFT JOIN [dbo].[book]       b ON bc.acc_no = CAST(b.acc_no AS VARCHAR)
        LEFT JOIN [dbo].[personal]   p ON bc.id_no  = p.id_no
        LEFT JOIN [dbo].[department] d ON p.dept_no = d.dept_no
        LEFT JOIN [dbo].[catagory]   c ON p.cat_no  = c.cat_no
        WHERE CAST(bc.due_date AS DATE) = CAST(GETDATE() AS DATE)
          AND p.e_mail IS NOT NULL AND p.e_mail != ''
        ORDER BY p.name, bc.due_date
    """)
    rows = cur.fetchall()
    conn.close()
    return _filter_prefixes(rows, prefixes)


def fetch_overdue_students(prefixes, active_settings=None):
    """Students whose book(s) are overdue — calculates fines directly.
    Uses catagory.fine_day * chargeable_days (excluding Sundays + holidays).
    Confirmed matching LIPS5 software calculation.
    active_settings: dict with keys 'first_year_seq', 'last_year_seq', 'email_domain'
    """
    conn = mssql_connect()
    if conn is None: return []
    cur = conn.cursor(as_dict=True)

    # ── Step 1: Fetch library control flags ──────────────────────────
    try:
        cur.execute("SELECT skip_sunday, skip_holiday FROM [dbo].[controls]")
        ctrl = cur.fetchone() or {}
    except Exception:
        ctrl = {}
    skip_sunday  = (ctrl.get("skip_sunday", 0) or 0) == 1
    skip_holiday = (ctrl.get("skip_holiday", 0) or 0) == 1

    # ── Step 2: Fetch ALL holidays from the database ─────────────────
    holidays = set()
    try:
        cur.execute("SELECT holy_date FROM [dbo].[holiday]")
        for h in cur.fetchall():
            hd = h.get("holy_date")
            if hd:
                if isinstance(hd, datetime.datetime):
                    holidays.add(hd.date())
                elif isinstance(hd, datetime.date):
                    holidays.add(hd)
    except Exception:
        pass

    # ── Step 3: Get the SQL Server date (source of truth) ────────────
    # NOTE: On Computer B (older TDS 7.0), dates may come back as strings
    # instead of datetime objects. We must handle both cases.
    try:
        cur.execute("SELECT CAST(GETDATE() AS DATE) AS server_today")
        server_today = cur.fetchone()["server_today"]
        if isinstance(server_today, datetime.datetime):
            server_today = server_today.date()
        elif isinstance(server_today, str):
            # Older TDS versions return dates as strings like "2026-03-30"
            server_today = datetime.datetime.strptime(server_today.strip()[:10], "%Y-%m-%d").date()
        # else: already a datetime.date — keep as-is
    except Exception:
        server_today = datetime.date.today()

    # ── Step 4: Main query — get fine_day from catagory directly ─────
    cur.execute("""
        SELECT
            bc.id_no         AS RollNumber,
            p.name           AS Name,
            ISNULL(p.e_mail,'') AS Email,
            ISNULL(d.dept_name,'') AS Department,
            b.title          AS BookTitle,
            ISNULL(b.author, '') AS Author,
            bc.acc_no        AS AccNo,
            bc.t_date        AS IssueDate,
            bc.due_date      AS DueDate,
            ISNULL(c.cat_name,'') AS Category,
            p.active_member  AS IsActiveMember,
            ISNULL(c.fine_day, 1) AS FinePerDay
        FROM [dbo].[book_circle] bc
        LEFT JOIN [dbo].[book]       b  ON bc.acc_no = CAST(b.acc_no AS VARCHAR)
        LEFT JOIN [dbo].[personal]   p  ON bc.id_no  = p.id_no
        LEFT JOIN [dbo].[department] d  ON p.dept_no = d.dept_no
        LEFT JOIN [dbo].[catagory]   c  ON p.cat_no  = c.cat_no
        WHERE bc.due_date < CAST(GETDATE() AS DATE)
          AND p.e_mail IS NOT NULL AND p.e_mail != ''
        ORDER BY p.name, bc.due_date
    """)
    rows = cur.fetchall()
    conn.close()

    # ── Step 5: Calculate chargeable days & fine for each row ────────
    for row in rows:
        due = row.get("DueDate")
        if isinstance(due, datetime.datetime):
            due = due.date()
        elif isinstance(due, str):
            # Older TDS versions return dates as strings
            try:
                due = datetime.datetime.strptime(due.strip()[:10], "%Y-%m-%d").date()
            except (ValueError, AttributeError):
                due = None
        elif not isinstance(due, datetime.date):
            due = None

        fine_per_day = row.get("FinePerDay", 1) or 1

        # Day-by-day loop from (due_date + 1) to server_today
        # This exactly matches the LIPS5 software calculation
        chargeable_days = 0
        if due and due < server_today:
            current_date = due + datetime.timedelta(days=1)
            while current_date <= server_today:
                skip = False
                if skip_sunday and current_date.weekday() == 6:
                    skip = True
                if skip_holiday and current_date in holidays:
                    skip = True
                if not skip:
                    chargeable_days += 1
                current_date += datetime.timedelta(days=1)

        row["OverdueDays"] = chargeable_days
        row["FineAmount"] = int(chargeable_days * fine_per_day)

        # Clean up temp field not needed downstream
        row.pop("FinePerDay", None)

    # Determine MemberType based on per-department settings
    for row in rows:
        row["MemberType"] = _classify_member_type(row, active_settings)

    # Filter by department-specific prefixes
    filtered_rows = []
    dept_configs = active_settings.get("depts", {})
    for r in rows:
        dept = r.get("Department", "")
        prefixes = dept_configs.get(dept, {}).get("prefixes", "")
        if not prefixes:
            filtered_rows.append(r)
        else:
            plist = [p.strip().upper() for p in prefixes.split(",") if p.strip()]
            if not plist or any(str(r["RollNumber"]).upper().startswith(p) for p in plist):
                filtered_rows.append(r)
    
    return filtered_rows


def _calc_fine_from_slab(fine_slabs, chargeable_days, cat_no):
    """Calculate fine using the fine_slab table when catagory.fine_day is 0/NULL."""
    if chargeable_days <= 0:
        return 0

    # Try to find a slab matching this category first, then fall back to any slab
    matching_slabs = [s for s in fine_slabs if s.get("cat_no") == cat_no] if cat_no else []
    slabs_to_use = matching_slabs if matching_slabs else fine_slabs

    if not slabs_to_use:
        # No slabs found at all — default to ₹1/day
        return chargeable_days

    total_fine = 0
    remaining_days = chargeable_days

    for slab in sorted(slabs_to_use, key=lambda s: s.get("day_from", 0) or 0):
        day_from = slab.get("day_from", 0) or 0
        day_to = slab.get("day_to", 999999) or 999999
        day_amt = slab.get("day_amt", 1) or 1

        if chargeable_days >= day_from:
            slab_days = min(chargeable_days, day_to) - day_from + 1
            if slab_days > 0:
                total_fine = slab_days * day_amt

    return total_fine if total_fine > 0 else chargeable_days


def _classify_member_type(row, active_settings=None):
    """Classify a member as Active Student, Passout Student, or Teacher.
    Uses active_settings for roll number range + email domain classification.
    
    Classification rules:
      - Teachers: identified by category keywords (TEACHER, STAFF, etc.)
      - Passout Students: roll number is None/empty OR falls OUTSIDE the
        configured range (first_seq → last_seq) in settings
      - Active Students: roll number falls WITHIN the configured range,
        active_member == 1, and email domain matches (if set)
    """
    cat_name = row.get("Category", "").upper()
    # Teacher classification by category name
    teacher_keywords = ['TEACHER', 'STAFF', 'LECTURER', 'PROF', 'RTEACHER', 'FACULTY']
    if any(kw in cat_name for kw in teacher_keywords):
        return "Teacher"

    # Student classification
    raw_roll = row.get("RollNumber")
    roll = str(raw_roll).upper() if raw_roll is not None else ""
    active_member = row.get("IsActiveMember", 0)
    email = str(row.get("Email", "")).lower().strip()

    # ── Rule: None / empty roll numbers are always Passout ────────────
    if raw_roll is None or roll in ("", "NONE", "NULL"):
        return "Passout Student"
    
    # ── Rule: None / empty emails are always Passout ────────────
    if not email or email == "none" or email == "null":
        return "Passout Student"

    # Get department-specific settings
    dept_name = row.get("Department", "")
    dept_configs = active_settings.get("depts", {}) if active_settings else {}
    dept_setting = dept_configs.get(dept_name, {})
    email_domain = active_settings.get("email_domain", "").lower().strip() if active_settings else ""

    if dept_setting:
        first_year_seq = dept_setting.get("first_seq", "").upper()
        last_year_seq = dept_setting.get("last_seq", "").upper()

        # Check if roll number falls within the configured active range
        in_range = True
        if first_year_seq and last_year_seq:
            in_range = False
            # Extract numeric parts
            roll_digits = "".join(c for c in roll if c.isdigit())
            first_digits = "".join(c for c in first_year_seq if c.isdigit())
            last_digits = "".join(c for c in last_year_seq if c.isdigit())
            
            try:
                if first_digits and last_digits and roll_digits:
                    # Truncate roll sequence to the length of the boundary for proper comparison
                    comp_len = len(first_digits)
                    # Use standard string slicing, ignoring Pyre false positives on generic slice
                    comp_roll_str = "".join(list(roll_digits)[:comp_len])
                    if comp_roll_str:
                        comp_roll = int(comp_roll_str)
                        v1 = int(first_digits)
                        v2 = int(last_digits)
                        
                        # Any students between the two years (e.g. 822 to 825)
                        in_range = min(v1, v2) <= comp_roll <= max(v1, v2)
            except ValueError:
                in_range = False

        # ── Rule: Outside the range → Passout ────────────────────────
        if not in_range:
            return "Passout Student"

    # ── Rule: Check email domain ────────────────────────
    if email_domain and '@' in email:
        has_active_domain = email.split('@')[-1].strip() == email_domain
        if not has_active_domain:
            return "Passout Student"
    elif email_domain and '@' not in email:
        # If an email domain is configured but the user's email has no '@', it's invalid
        return "Passout Student"

    # Fallback to active member flag if all checks pass
    if active_member == 1:
        return "Active Student"
    
    return "Passout Student"


def _filter_prefixes(rows, prefixes):
    """Keep only rows whose RollNumber starts with one of the allowed prefixes."""
    if not prefixes:
        return rows
    return [r for r in rows if any(str(r["RollNumber"]).upper().startswith(p.upper()) for p in prefixes)]


# ══════════════════════════════════════════════════════════════════════════
#  EMAIL helper
# ══════════════════════════════════════════════════════════════════════════
def send_email(to_addr, subject, html_body):
    msg = MIMEMultipart("alternative")
    msg["From"] = f"HiTECH Library <{GMAIL_USER}>"
    msg["To"] = to_addr
    msg["Subject"] = subject
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    with smtplib.SMTP_SSL("smtp.gmail.com", 465, timeout=20) as s:
        s.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        s.sendmail(GMAIL_USER, to_addr, msg.as_string())


def build_reminder_html(name, books):
    rows = ""
    for b in books:
        due = b["DueDate"]
        if isinstance(due, datetime.datetime):
            due = due.strftime("%d %b %Y")
        rows += f"""<tr>
            <td style="padding:10px;border-bottom:1px solid #e2e8f0;font-size:13px">{b['BookTitle']}</td>
            <td style="padding:10px;border-bottom:1px solid #e2e8f0;font-size:13px;text-align:center">{b['AccNo']}</td>
            <td style="padding:10px;border-bottom:1px solid #e2e8f0;font-size:13px;text-align:center">{due}</td>
        </tr>"""
    return f"""
    <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:520px;margin:0 auto;background:#f8fafc;border-radius:16px;overflow:hidden;border:1px solid #e2e8f0">
        <div style="background:linear-gradient(135deg,#6366f1,#7c3aed);padding:24px;text-align:center">
            <h1 style="color:#fff;margin:0;font-size:20px;font-weight:800">📚 HiTECH Library</h1>
            <p style="color:rgba(255,255,255,.85);margin:4px 0 0;font-size:12px">Hindustan Institute of Technology</p>
        </div>
        <div style="padding:24px">
            <h3 style="color:#f59e0b;margin:0 0 12px">⏰ Due Date Reminder</h3>
            <p style="color:#475569;margin:0 0 16px">Hi <strong>{name}</strong>,</p>
            <p style="color:#64748b;font-size:14px;margin:0 0 12px">The following book(s) are due <strong>tomorrow</strong>. Please return them to avoid fines.</p>
            <table style="width:100%;border-collapse:collapse;margin-bottom:16px">
                <thead><tr style="background:#f8fafc">
                    <th style="padding:10px;text-align:left;font-size:12px;text-transform:uppercase;color:#64748b">Book</th>
                    <th style="padding:10px;text-align:center;font-size:12px;text-transform:uppercase;color:#64748b">Acc No</th>
                    <th style="padding:10px;text-align:center;font-size:12px;text-transform:uppercase;color:#64748b">Due Date</th>
                </tr></thead>
                <tbody>{rows}</tbody>
            </table>
            <p style="color:#64748b;font-size:14px">A fine of ₹1 per day will be charged after the due date (holidays excluded).</p>
        </div>
        <div style="background:#f1f5f9;padding:16px 24px;text-align:center;border-top:1px solid #e2e8f0">
            <p style="color:#94a3b8;font-size:11px;margin:0">© 2026 HiTECH Library. All rights reserved.</p>
        </div>
    </div>"""


def build_overdue_html(name, books, total_fine):
    rows = ""
    for b in books:
        due = b["DueDate"]
        if isinstance(due, datetime.datetime):
            due = due.strftime("%d %b %Y")
        rows += f"""<tr>
            <td style="padding:10px;border-bottom:1px solid #f1f5f9;font-size:13px;color:#334155">{b['BookTitle']}</td>
            <td style="padding:10px;border-bottom:1px solid #f1f5f9;font-size:13px;text-align:center;color:#64748b">{due}</td>
            <td style="padding:10px;border-bottom:1px solid #f1f5f9;font-size:13px;text-align:center;color:#dc2626;font-weight:700">{b.get('OverdueDays',0)} days</td>
            <td style="padding:10px;border-bottom:1px solid #f1f5f9;font-size:13px;text-align:center;color:#dc2626;font-weight:700">₹{b.get('FineAmount',0)}</td>
        </tr>"""
    return f"""
    <div style="font-family:'Segoe UI',Arial,sans-serif;max-width:520px;margin:0 auto;background:#f8fafc;border-radius:16px;overflow:hidden;border:1px solid #e2e8f0">
        <div style="background:linear-gradient(135deg,#6366f1,#7c3aed);padding:24px;text-align:center">
            <h1 style="color:#fff;margin:0;font-size:20px;font-weight:800">📚 HiTECH Library</h1>
            <p style="color:rgba(255,255,255,.85);margin:4px 0 0;font-size:12px">Hindustan Institute of Technology</p>
        </div>
        <div style="padding:24px">
            <h3 style="color:#dc2626;margin:0 0 12px">⚠️ Overdue Books Alert</h3>
            <p style="color:#475569;margin:0 0 16px">Hi <strong>{name}</strong>,</p>
            <p style="color:#64748b;font-size:14px;margin:0 0 16px">You have <strong>{len(books)} overdue book(s)</strong>. Please return them immediately to stop accumulating fines.</p>
            <table style="width:100%;border-collapse:collapse;margin-bottom:16px">
                <thead><tr style="background:#f8fafc">
                    <th style="padding:10px;text-align:left;font-size:12px;text-transform:uppercase;color:#64748b">Book</th>
                    <th style="padding:10px;text-align:center;font-size:12px;text-transform:uppercase;color:#64748b">Due Date</th>
                    <th style="padding:10px;text-align:center;font-size:12px;text-transform:uppercase;color:#64748b">Overdue</th>
                    <th style="padding:10px;text-align:center;font-size:12px;text-transform:uppercase;color:#64748b">Fine</th>
                </tr></thead>
                <tbody>{rows}</tbody>
            </table>
            <div style="background:#fef2f2;border:2px solid #fecaca;border-radius:12px;padding:16px;text-align:center">
                <p style="margin:0;color:#7f1d1d;font-size:13px">Total Fine Amount</p>
                <p style="margin:4px 0 0;color:#dc2626;font-size:28px;font-weight:900">₹{total_fine}</p>
            </div>
            <p style="color:#94a3b8;font-size:12px;margin:16px 0 0;text-align:center">* Holidays are excluded from fine calculation.</p>
        </div>
        <div style="background:#f1f5f9;padding:16px 24px;text-align:center;border-top:1px solid #e2e8f0">
            <p style="color:#94a3b8;font-size:11px;margin:0">© 2026 HiTECH Library. All rights reserved.</p>
        </div>
    </div>"""


# ══════════════════════════════════════════════════════════════════════════
#  GROUP ROWS BY STUDENT
# ══════════════════════════════════════════════════════════════════════════
def group_by_student(rows):
    """Returns dict: RollNumber -> {Name, Email, Department, Category, MemberType, books: []}"""
    groups = {}
    for r in rows:
        key = r["RollNumber"]
        if key not in groups:
            groups[key] = {
                "RollNumber": key,
                "Name": r.get("Name", ""),
                "Email": r.get("Email", ""),
                "Department": r.get("Department", ""),
                "Category": r.get("Category", ""),
                "MemberType": r.get("MemberType", "Active Student"),
                "books": [],
            }
        groups[key]["books"].append(r)
    return groups


# ══════════════════════════════════════════════════════════════════════════
#  CUSTOMTKINTER APP
# ══════════════════════════════════════════════════════════════════════════
ctk.set_appearance_mode("light")
ctk.set_default_color_theme("blue")

FONT_TITLE = ("Segoe UI", 22, "bold")
FONT_HEADING = ("Segoe UI", 15, "bold")
FONT_BODY = ("Segoe UI", 13)
FONT_SMALL = ("Segoe UI", 11)
FONT_BADGE = ("Segoe UI", 11, "bold")

PRIMARY = "#6366f1"
PRIMARY_HOVER = "#4f46e5"
DANGER = "#dc2626"
SUCCESS = "#16a34a"
WARNING = "#f59e0b"
SURFACE = "#ffffff"
BG = "#f1f5f9"
BORDER = "#e2e8f0"
TEXT_PRIMARY = "#1e293b"
TEXT_SECONDARY = "#64748b"


class App(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("📚 HiTECH Library — Librarian Email Manager")
        self.geometry("1180x740")
        self.minsize(960, 600)
        self.configure(fg_color=BG)

        init_local_db()

        # ─── data stores ──────────────────────────────────────────────
        self.reminder_data = {}   # grouped by student
        self.today_data = {}
        self.overdue_data = {}
        self.reminder_checks = {}  # roll -> BooleanVar
        self.today_checks = {}
        self.overdue_checks = {}
        self.reminder_widgets = []
        self.today_widgets = []
        self.overdue_widgets = []
        
        self.history_map = get_all_sent_history_counts()
        self.search_var = ctk.StringVar()
        self.search_var.trace_add("write", lambda *args: self._trigger_search())

        self._build_ui()
        self._load_prefixes()
        self._load_active_settings()
        self._check_connection()  # test DB on startup

    # ─── UI SKELETON ──────────────────────────────────────────────────
    def _build_ui(self):
        # top bar
        top = ctk.CTkFrame(self, fg_color=PRIMARY, corner_radius=0, height=56)
        top.pack(fill="x")
        top.pack_propagate(False)
        ctk.CTkLabel(top, text="  HiTECH Library  —  Email Manager",
                     font=("Segoe UI", 18, "bold"), text_color="#ffffff").pack(side="left", padx=20)

        # connection status badge (right side)
        self.conn_badge = ctk.CTkLabel(
            top, text="  Checking...  ",
            font=("Segoe UI", 12, "bold"), text_color="#ffffff",
            fg_color="#6b7280", corner_radius=8,
            width=180, height=30
        )
        self.conn_badge.pack(side="right", padx=20, pady=13)

        self.status_lbl = ctk.CTkLabel(top, text="", font=FONT_SMALL, text_color="#c7d2fe")
        self.status_lbl.pack(side="right", padx=(0, 8))
        
        # Search Entry (right side, before status)
        self.search_entry = ctk.CTkEntry(
            top, textvariable=self.search_var, placeholder_text="🔍 Search Name / Roll...",
            width=200, height=30, fg_color="#4f46e5", border_color="#818cf8", text_color="#ffffff"
        )
        self.search_entry.pack(side="right", padx=(0, 15), pady=13)
        
        # Export Buttons Container (right side)
        export_frame = ctk.CTkFrame(top, fg_color="transparent")
        export_frame.pack(side="right", padx=(0, 15), pady=10)

        ctk.CTkButton(
            export_frame, text="📥  Export Data", font=FONT_SMALL, fg_color="#10b981", hover_color="#059669",
            width=100, height=32, command=lambda: self._export_to_excel("all_cols")
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            export_frame, text="👥  Export Students", font=FONT_SMALL, fg_color="#10b981", hover_color="#059669",
            width=110, height=32, command=lambda: self._export_to_excel("students_only")
        ).pack(side="left", padx=2)

        ctk.CTkButton(
            export_frame, text="📄  Export Details", font=FONT_SMALL, fg_color="#10b981", hover_color="#059669",
            width=110, height=32, command=lambda: self._export_to_excel("detailed")
        ).pack(side="left", padx=2)

        # ── CUSTOM TAB BAR ──────────────────────────────────────────
        tab_bar = ctk.CTkFrame(self, fg_color=SURFACE, corner_radius=0, height=46)
        tab_bar.pack(fill="x", padx=0, pady=0)
        tab_bar.pack_propagate(False)

        inner_bar = ctk.CTkFrame(tab_bar, fg_color="transparent")
        inner_bar.pack(expand=True)

        self._tab_buttons = {}
        tab_defs = [
            ("reminder", "Reminder (Tmrw)"),
            ("today",    "Today's Due"),
            ("overdue",  "Overdue & Fines"),
            ("history",  "History"),
            ("settings", "Settings"),
        ]
        for key, label in tab_defs:
            btn = ctk.CTkButton(
                inner_bar, text=label,
                font=("Segoe UI", 13, "bold"), width=170, height=36, corner_radius=8,
                fg_color="transparent", text_color=TEXT_SECONDARY,
                hover_color="#e0e7ff",
                command=lambda k=key: self._switch_tab(k),
            )
            btn.pack(side="left", padx=4, pady=5)
            self._tab_buttons[key] = btn

        # tab separator
        ctk.CTkFrame(self, fg_color=BORDER, height=1, corner_radius=0).pack(fill="x")

        # ── TAB FRAMES ──────────────────────────────────────────────
        self._tab_container = ctk.CTkFrame(self, fg_color=BG, corner_radius=0)
        self._tab_container.pack(fill="both", expand=True)

        self.tab_reminder = ctk.CTkFrame(self._tab_container, fg_color=BG, corner_radius=0)
        self.tab_today    = ctk.CTkFrame(self._tab_container, fg_color=BG, corner_radius=0)
        self.tab_overdue  = ctk.CTkFrame(self._tab_container, fg_color=BG, corner_radius=0)
        self.tab_history  = ctk.CTkFrame(self._tab_container, fg_color=BG, corner_radius=0)
        self.tab_settings = ctk.CTkFrame(self._tab_container, fg_color=BG, corner_radius=0)

        self._tab_frames = {
            "reminder": self.tab_reminder,
            "today":    self.tab_today,
            "overdue":  self.tab_overdue,
            "history":  self.tab_history,
            "settings": self.tab_settings,
        }

        self._build_reminder_tab()
        self._build_today_tab()
        self._build_overdue_tab()
        self._build_history_tab()
        self._build_settings_tab()

        # show first tab
        self._switch_tab("reminder")

    def _switch_tab(self, key):
        """Show the selected tab frame, highlight its button."""
        self.current_tab = key
        for k, frame in self._tab_frames.items():
            frame.pack_forget()
        self._tab_frames[key].pack(in_=self._tab_container, fill="both", expand=True)

        for k, btn in self._tab_buttons.items():
            if k == key:
                btn.configure(fg_color=PRIMARY, text_color="#ffffff", hover_color=PRIMARY_HOVER)
            else:
                btn.configure(fg_color="transparent", text_color=TEXT_SECONDARY, hover_color="#e0e7ff")
        
        self._trigger_search()

    # ═══════════════════════════════════════════════════════════════════
    #  REMINDER TAB
    # ═══════════════════════════════════════════════════════════════════
    def _build_reminder_tab(self):
        tab = self.tab_reminder

        # toolbar
        bar = ctk.CTkFrame(tab, fg_color="transparent")
        bar.pack(fill="x", padx=8, pady=(8, 4))

        ctk.CTkButton(bar, text="🔄  Fetch Data", font=FONT_BODY, fg_color=PRIMARY,
                      hover_color=PRIMARY_HOVER, command=self._fetch_reminder, width=140).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="☑️  Select All", font=FONT_BODY, fg_color="#6b7280",
                      hover_color="#4b5563", command=lambda: self._select_all("reminder"), width=120).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="☐  Deselect All", font=FONT_BODY, fg_color="#6b7280",
                      hover_color="#4b5563", command=lambda: self._deselect_all("reminder"), width=120).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="📤  Send Selected", font=FONT_BODY, fg_color=SUCCESS,
                      hover_color="#15803d", command=lambda: self._send_selected("reminder"), width=140).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="📤  Send All", font=FONT_BODY, fg_color=WARNING,
                      hover_color="#d97706", text_color="#1e293b",
                      command=lambda: self._send_all("reminder"), width=120).pack(side="left", padx=4)

        self.reminder_count_lbl = ctk.CTkLabel(bar, text="0 students", font=FONT_BADGE, text_color=TEXT_SECONDARY)
        self.reminder_count_lbl.pack(side="right", padx=12)

        # scrollable list
        self.reminder_scroll = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=8)
        self.reminder_scroll.pack(fill="both", expand=True, padx=8, pady=4)

        self.reminder_empty = ctk.CTkLabel(self.reminder_scroll, text='Click "Fetch Data" to load students with books due tomorrow.',
                                           font=FONT_BODY, text_color=TEXT_SECONDARY)
        self.reminder_empty.pack(pady=40)

    # ═══════════════════════════════════════════════════════════════════
    #  TODAY TAB
    # ═══════════════════════════════════════════════════════════════════
    def _build_today_tab(self):
        tab = self.tab_today

        # toolbar
        bar = ctk.CTkFrame(tab, fg_color="transparent")
        bar.pack(fill="x", padx=8, pady=(8, 4))

        ctk.CTkButton(bar, text="🔄  Fetch Data", font=FONT_BODY, fg_color=PRIMARY,
                      hover_color=PRIMARY_HOVER, command=self._fetch_today, width=140).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="☑️  Select All", font=FONT_BODY, fg_color="#6b7280",
                      hover_color="#4b5563", command=lambda: self._select_all("today"), width=120).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="☐  Deselect All", font=FONT_BODY, fg_color="#6b7280",
                      hover_color="#4b5563", command=lambda: self._deselect_all("today"), width=120).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="📤  Send Selected", font=FONT_BODY, fg_color=SUCCESS,
                      hover_color="#15803d", command=lambda: self._send_selected("today"), width=140).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="📤  Send All", font=FONT_BODY, fg_color=WARNING,
                      hover_color="#d97706", text_color="#1e293b",
                      command=lambda: self._send_all("today"), width=120).pack(side="left", padx=4)

        self.today_count_lbl = ctk.CTkLabel(bar, text="0 students", font=FONT_BADGE, text_color=TEXT_SECONDARY)
        self.today_count_lbl.pack(side="right", padx=12)

        # scrollable list
        self.today_scroll = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=8)
        self.today_scroll.pack(fill="both", expand=True, padx=8, pady=4)

        self.today_empty = ctk.CTkLabel(self.today_scroll, text='Click "Fetch Data" to load students with books due today.',
                                        font=FONT_BODY, text_color=TEXT_SECONDARY)
        self.today_empty.pack(pady=40)

    # ═══════════════════════════════════════════════════════════════════
    #  OVERDUE TAB
    # ═══════════════════════════════════════════════════════════════════
    def _build_overdue_tab(self):
        tab = self.tab_overdue

        # toolbar
        bar = ctk.CTkFrame(tab, fg_color="transparent")
        bar.pack(fill="x", padx=8, pady=(8, 4))

        ctk.CTkButton(bar, text="🔄  Fetch Data", font=FONT_BODY, fg_color=PRIMARY,
                      hover_color=PRIMARY_HOVER, command=self._fetch_overdue, width=140).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="☑️  Select All", font=FONT_BODY, fg_color="#6b7280",
                      hover_color="#4b5563", command=lambda: self._select_all("overdue"), width=120).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="☐  Deselect All", font=FONT_BODY, fg_color="#6b7280",
                      hover_color="#4b5563", command=lambda: self._deselect_all("overdue"), width=120).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="📤  Send Selected", font=FONT_BODY, fg_color=SUCCESS,
                      hover_color="#15803d", command=lambda: self._send_selected("overdue"), width=140).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="📤  Send All", font=FONT_BODY, fg_color=WARNING,
                      hover_color="#d97706", text_color="#1e293b",
                      command=lambda: self._send_all("overdue"), width=120).pack(side="left", padx=4)

        self.overdue_count_lbl = ctk.CTkLabel(bar, text="0 students", font=FONT_BADGE, text_color=TEXT_SECONDARY)
        self.overdue_count_lbl.pack(side="right", padx=12)

        # sort bar
        sort_bar = ctk.CTkFrame(tab, fg_color="transparent")
        sort_bar.pack(fill="x", padx=8, pady=(0, 4))

        ctk.CTkLabel(sort_bar, text="Sort by:", font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(side="left", padx=(4, 8))
        self.overdue_sort = ctk.CTkComboBox(sort_bar, values=[
            "Name (A-Z)", "Days Overdue (High→Low)", "Fine Amount (High→Low)",
            "Year/Batch (Newest First)", "Department (A-Z)", "Book Title (A-Z)"
        ], font=FONT_SMALL, width=220, command=self._sort_overdue_display)
        self.overdue_sort.set("Days Overdue (High→Low)")
        self.overdue_sort.pack(side="left", padx=4)

        # range filter
        ctk.CTkLabel(sort_bar, text="  Days range:", font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(side="left", padx=(12, 4))
        self.overdue_min_days = ctk.CTkEntry(sort_bar, width=50, font=FONT_SMALL, placeholder_text="Min")
        self.overdue_min_days.pack(side="left", padx=2)
        ctk.CTkLabel(sort_bar, text="–", font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(side="left")
        self.overdue_max_days = ctk.CTkEntry(sort_bar, width=50, font=FONT_SMALL, placeholder_text="Max")
        self.overdue_max_days.pack(side="left", padx=2)

        # dept filter
        ctk.CTkLabel(sort_bar, text="  Dept:", font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(side="left", padx=(12, 4))
        self.overdue_dept_filter = ctk.CTkComboBox(sort_bar, values=["All"], font=FONT_SMALL, width=160)
        self.overdue_dept_filter.set("All")
        self.overdue_dept_filter.pack(side="left", padx=2)

        ctk.CTkButton(sort_bar, text="Apply", font=FONT_SMALL, fg_color=PRIMARY,
                      hover_color=PRIMARY_HOVER, width=70,
                      command=self._sort_overdue_display).pack(side="left", padx=8)

        # ── SEGREGATION TABS ─────────────────────────────────────────
        self.overdue_tabs = ctk.CTkTabview(tab, fg_color=BG, segmented_button_fg_color=SURFACE,
                                           segmented_button_selected_color=PRIMARY,
                                           segmented_button_selected_hover_color=PRIMARY_HOVER,
                                           command=self._on_overdue_tab_change)
        self.overdue_tabs.pack(fill="both", expand=True, padx=8, pady=4)

        self.tab_active_students = self.overdue_tabs.add("Active Students")
        self.tab_passout_students = self.overdue_tabs.add("Passout Students")
        self.tab_teachers = self.overdue_tabs.add("Teachers")

        # scrollable lists for each tab
        self.active_scroll = ctk.CTkScrollableFrame(self.tab_active_students, fg_color=BG, corner_radius=0)
        self.active_scroll.pack(fill="both", expand=True)

        self.passout_scroll = ctk.CTkScrollableFrame(self.tab_passout_students, fg_color=BG, corner_radius=0)
        self.passout_scroll.pack(fill="both", expand=True)

        self.teacher_scroll = ctk.CTkScrollableFrame(self.tab_teachers, fg_color=BG, corner_radius=0)
        self.teacher_scroll.pack(fill="both", expand=True)

        # Empty labels (placeholders)
        self.active_empty_lbl = ctk.CTkLabel(self.active_scroll, text='No active students found.', font=FONT_BODY, text_color=TEXT_SECONDARY)
        self.active_empty_lbl.pack(pady=40)
        
        self.passout_empty_lbl = ctk.CTkLabel(self.passout_scroll, text='No passout students found.', font=FONT_BODY, text_color=TEXT_SECONDARY)
        self.passout_empty_lbl.pack(pady=40)
        
        self.teacher_empty_lbl = ctk.CTkLabel(self.teacher_scroll, text='No teachers found.', font=FONT_BODY, text_color=TEXT_SECONDARY)
        self.teacher_empty_lbl.pack(pady=40)

    # ═══════════════════════════════════════════════════════════════════
    #  HISTORY TAB
    # ═══════════════════════════════════════════════════════════════════
    def _build_history_tab(self):
        tab = self.tab_history

        bar = ctk.CTkFrame(tab, fg_color="transparent")
        bar.pack(fill="x", padx=8, pady=(8, 4))
        ctk.CTkButton(bar, text="🔄  Refresh", font=FONT_BODY, fg_color=PRIMARY,
                      hover_color=PRIMARY_HOVER, command=self._load_history, width=120).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="🗑️  Delete Selected", font=FONT_BODY, fg_color=DANGER,
                      hover_color="#b91c1c", command=self._delete_selected_history, width=150).pack(side="left", padx=4)
        ctk.CTkButton(bar, text="☑️  Select All", font=FONT_BODY, fg_color="#6b7280",
                      hover_color="#4b5563", command=self._select_all_history, width=120).pack(side="left", padx=4)

        self.history_count_lbl = ctk.CTkLabel(bar, text="", font=FONT_BADGE, text_color=TEXT_SECONDARY)
        self.history_count_lbl.pack(side="right", padx=12)

        self.history_scroll = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=8)
        self.history_scroll.pack(fill="both", expand=True, padx=8, pady=4)

        self.history_checks = {}
        self.history_widgets = []

    # ═══════════════════════════════════════════════════════════════════
    #  SETTINGS TAB
    # ═══════════════════════════════════════════════════════════════════
    def _build_settings_tab(self):
        tab = self.tab_settings

        # Main scrollable container for all settings
        settings_scroll = ctk.CTkScrollableFrame(tab, fg_color=BG, corner_radius=0)
        settings_scroll.pack(fill="both", expand=True, padx=0, pady=0)

        # ─── SECTION 1: Active Student Classification ──────────────────
        active_wrapper = ctk.CTkFrame(settings_scroll, fg_color=SURFACE, corner_radius=12,
                                      border_width=1, border_color=BORDER)
        active_wrapper.pack(fill="x", padx=16, pady=(16, 8))

        ctk.CTkLabel(active_wrapper, text="🎓  Active Student Classification",
                     font=FONT_HEADING, text_color=TEXT_PRIMARY).pack(anchor="w", padx=20, pady=(20, 4))
        ctk.CTkLabel(active_wrapper,
                     text="Define the roll number range and email domain for currently studying (active) students. "
                          "Members with roll numbers between First Year and Last Year sequences, "
                          "who have active_member = 1 AND the matching email domain will be classified as Active Students. "
                          "All others (except teachers) will be considered Passed Out Students.",
                     font=FONT_SMALL, text_color=TEXT_SECONDARY, wraplength=700, justify="left"
                     ).pack(anchor="w", padx=20, pady=(0, 12))

        # Roll number range row
        range_frame = ctk.CTkFrame(active_wrapper, fg_color=BG, corner_radius=8)
        range_frame.pack(fill="x", padx=20, pady=4)

        range_inner = ctk.CTkFrame(range_frame, fg_color="transparent")
        range_inner.pack(fill="x", padx=12, pady=12)

        ctk.CTkLabel(range_inner, text="First Year Seq Roll No:", font=FONT_BODY,
                     text_color=TEXT_PRIMARY).pack(side="left", padx=(0, 8))
        self.first_year_entry = ctk.CTkEntry(range_inner, width=180, font=FONT_BODY,
                                             placeholder_text="e.g. 22CSR001")
        self.first_year_entry.pack(side="left", padx=(0, 20))

        ctk.CTkLabel(range_inner, text="Last Year Seq Roll No:", font=FONT_BODY,
                     text_color=TEXT_PRIMARY).pack(side="left", padx=(0, 8))
        self.last_year_entry = ctk.CTkEntry(range_inner, width=180, font=FONT_BODY,
                                            placeholder_text="e.g. 25CSR999")
        self.last_year_entry.pack(side="left", padx=(0, 8))

        # Email domain row
        domain_frame = ctk.CTkFrame(active_wrapper, fg_color=BG, corner_radius=8)
        domain_frame.pack(fill="x", padx=20, pady=4)

        domain_inner = ctk.CTkFrame(domain_frame, fg_color="transparent")
        domain_inner.pack(fill="x", padx=12, pady=12)

        ctk.CTkLabel(domain_inner, text="Active Email Domain (after @):", font=FONT_BODY,
                     text_color=TEXT_PRIMARY).pack(side="left", padx=(0, 8))
        self.email_domain_entry = ctk.CTkEntry(domain_inner, width=280, font=FONT_BODY,
                                               placeholder_text="e.g. student.hits.ac.in")
        self.email_domain_entry.pack(side="left", padx=(0, 8))
        ctk.CTkLabel(domain_inner,
                     text="Members with this domain → Active.  Others → Passout.",
                     font=FONT_SMALL, text_color=TEXT_SECONDARY).pack(side="left", padx=(8, 0))

        # Save button for active settings
        active_btn_row = ctk.CTkFrame(active_wrapper, fg_color="transparent")
        active_btn_row.pack(fill="x", padx=20, pady=(4, 16))
        ctk.CTkButton(active_btn_row, text="💾  Save Active Settings", font=FONT_BODY, fg_color=SUCCESS,
                      hover_color="#15803d", command=self._save_active_settings, width=200).pack(side="left", padx=4)

        # ─── SECTION 2: Departments (Auto-fetched) ─────────────────────
        dept_wrapper = ctk.CTkFrame(settings_scroll, fg_color=SURFACE, corner_radius=12,
                                    border_width=1, border_color=BORDER)
        dept_wrapper.pack(fill="x", padx=16, pady=8)

        dept_title_row = ctk.CTkFrame(dept_wrapper, fg_color="transparent")
        dept_title_row.pack(fill="x", padx=20, pady=(20, 4))
        ctk.CTkLabel(dept_title_row, text="🏫  Departments (Auto-fetched from DB)",
                     font=FONT_HEADING, text_color=TEXT_PRIMARY).pack(side="left")
        ctk.CTkButton(dept_title_row, text="🔄 Refresh", font=FONT_SMALL, fg_color=PRIMARY,
                      hover_color=PRIMARY_HOVER, width=100, height=28,
                      command=self._fetch_departments_ui).pack(side="right")

        ctk.CTkLabel(dept_wrapper,
                     text="These departments are auto-fetched from the database. Use them to understand your member structure.",
                     font=FONT_SMALL, text_color=TEXT_SECONDARY, wraplength=700, justify="left"
                     ).pack(anchor="w", padx=20, pady=(0, 8))

        self.dept_list_frame = ctk.CTkFrame(dept_wrapper, fg_color=BG, corner_radius=8)
        self.dept_list_frame.pack(fill="x", padx=20, pady=(0, 16))

        self.dept_loading_lbl = ctk.CTkLabel(self.dept_list_frame,
                                             text="Click Refresh to load departments...",
                                             font=FONT_BODY, text_color=TEXT_SECONDARY)
        self.dept_loading_lbl.pack(pady=12)

        # ─── SECTION 3: Reg No. Prefix Sequences ──────────────────────
        prefix_wrapper = ctk.CTkFrame(settings_scroll, fg_color=SURFACE, corner_radius=12,
                                      border_width=1, border_color=BORDER)
        prefix_wrapper.pack(fill="x", padx=16, pady=8)

        ctk.CTkLabel(prefix_wrapper, text="⚙️  Reg No. Prefix Sequences",
                     font=FONT_HEADING, text_color=TEXT_PRIMARY).pack(anchor="w", padx=20, pady=(20, 4))
        ctk.CTkLabel(prefix_wrapper,
                     text="Only students whose Roll Number starts with one of these prefixes will appear in Reminder/Today tabs. "
                          "This helps filter to only active batch students.",
                     font=FONT_SMALL, text_color=TEXT_SECONDARY, wraplength=700, justify="left"
                     ).pack(anchor="w", padx=20, pady=(0, 12))

        # prefix list area
        self.prefix_frame = ctk.CTkFrame(prefix_wrapper, fg_color=BG, corner_radius=8)
        self.prefix_frame.pack(fill="x", padx=20, pady=4)

        self.prefix_entries = []
        self.prefix_container = ctk.CTkFrame(self.prefix_frame, fg_color="transparent")
        self.prefix_container.pack(fill="x", padx=8, pady=8)

        btn_row = ctk.CTkFrame(prefix_wrapper, fg_color="transparent")
        btn_row.pack(fill="x", padx=20, pady=(4, 16))
        ctk.CTkButton(btn_row, text="➕  Add Sequence", font=FONT_BODY, fg_color=PRIMARY,
                      hover_color=PRIMARY_HOVER, command=self._add_prefix_row, width=160).pack(side="left", padx=4)
        ctk.CTkButton(btn_row, text="💾  Save Prefixes", font=FONT_BODY, fg_color=SUCCESS,
                      hover_color="#15803d", command=self._save_prefixes, width=140).pack(side="left", padx=4)

        # ─── SECTION 4: Email Configuration ────────────────────────────
        smtp_wrapper = ctk.CTkFrame(settings_scroll, fg_color=SURFACE, corner_radius=12,
                                    border_width=1, border_color=BORDER)
        smtp_wrapper.pack(fill="x", padx=16, pady=(8, 16))

        ctk.CTkLabel(smtp_wrapper, text="📧  Email Configuration (from .env)",
                     font=FONT_HEADING, text_color=TEXT_PRIMARY).pack(anchor="w", padx=20, pady=(20, 8))

        info_frame = ctk.CTkFrame(smtp_wrapper, fg_color=BG, corner_radius=8)
        info_frame.pack(fill="x", padx=20, pady=(0, 16))
        ctk.CTkLabel(info_frame, text=f"GMAIL_USER:  {GMAIL_USER or '(not set)'}",
                     font=FONT_BODY, text_color=TEXT_PRIMARY).pack(anchor="w", padx=16, pady=(12, 2))
        ctk.CTkLabel(info_frame, text=f"GMAIL_APP_PASSWORD:  {'••••••••' if GMAIL_APP_PASSWORD else '(not set)'}",
                     font=FONT_BODY, text_color=TEXT_PRIMARY).pack(anchor="w", padx=16, pady=(2, 4))
        ctk.CTkLabel(info_frame, text=f"DB_SERVER:  {DB_SERVER}   |   DB_NAME:  {DB_NAME}",
                     font=FONT_BODY, text_color=TEXT_PRIMARY).pack(anchor="w", padx=16, pady=(2, 12))

    # ─── prefix helpers ───────────────────────────────────────────────
    def _add_prefix_row(self, value=""):
        row = ctk.CTkFrame(self.prefix_container, fg_color="transparent")
        row.pack(fill="x", pady=2)
        entry = ctk.CTkEntry(row, width=200, font=FONT_BODY, placeholder_text="e.g. 21CSR, 22CSD")
        entry.pack(side="left", padx=4)
        if value:
            entry.insert(0, value)
        remove_btn = ctk.CTkButton(row, text="✕", width=32, fg_color=DANGER,
                                   hover_color="#b91c1c", font=FONT_BADGE,
                                   command=lambda r=row, e=entry: self._remove_prefix_row(r, e))
        remove_btn.pack(side="left", padx=4)
        self.prefix_entries.append((row, entry))

    def _remove_prefix_row(self, row, entry):
        self.prefix_entries = [(r, e) for r, e in self.prefix_entries if r != row]
        row.destroy()

    def _save_prefixes(self):
        prefixes = [e.get().strip() for _, e in self.prefix_entries if e.get().strip()]
        set_setting("prefixes", json.dumps(prefixes))
        self._set_status(f"✅ Saved {len(prefixes)} prefix(es)")
        messagebox.showinfo("Saved", f"Saved {len(prefixes)} prefix sequence(s).")

    def _load_prefixes(self):
        raw = get_setting("prefixes", "[]")
        try:
            prefixes = json.loads(raw)
        except Exception:
            prefixes = []
        if not prefixes:
            self._add_prefix_row("")
        else:
            for p in prefixes:
                self._add_prefix_row(p)

    def _get_prefixes(self):
        raw = get_setting("prefixes", "[]")
        try:
            return json.loads(raw)
        except Exception:
            return []

    # ─── Active Settings helpers ──────────────────────────────────────
    def _save_active_settings(self):
        # Gather dept specific settings
        dept_configs = {}
        if hasattr(self, 'dept_entry_widgets'):
            for dname, widgets in self.dept_entry_widgets.items():
                f_e, l_e, p_e = widgets
                dept_configs[dname] = {
                    "first_seq": f_e.get().strip(),
                    "last_seq": l_e.get().strip(),
                    "prefixes": p_e.get().strip()
                }

        settings = {
            "email_domain": self.email_domain_entry.get().strip(),
            "depts": dept_configs
        }
        set_setting("active_settings", json.dumps(settings))
        self._set_status("✅ Dept-wise settings saved")
        messagebox.showinfo("Saved", "Department-wise classification settings saved successfully to local SQLite database.")

    def _load_active_settings(self):
        raw = get_setting("active_settings", "{}")
        try:
            settings = json.loads(raw)
        except Exception:
            settings = {}
        if settings.get("first_year_seq"):
            self.first_year_entry.insert(0, settings["first_year_seq"])
        if settings.get("last_year_seq"):
            self.last_year_entry.insert(0, settings["last_year_seq"])
        if settings.get("email_domain"):
            self.email_domain_entry.insert(0, settings["email_domain"])

    def _get_active_settings(self):
        raw = get_setting("active_settings", "{}")
        try:
            return json.loads(raw)
        except Exception:
            return {}

    # ─── Department fetch for settings UI ─────────────────────────────
    def _fetch_departments_ui(self):
        self._set_status("⏳ Fetching departments...")
        threading.Thread(target=self._fetch_departments_thread, daemon=True).start()

    def _fetch_departments_thread(self):
        try:
            depts = fetch_departments()
            self.after(0, lambda: self._render_departments(depts))
            self.after(0, lambda: self._update_conn_badge(True))
            self.after(0, lambda: self._set_status(f"✅ Loaded {len(depts)} departments"))
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.after(0, lambda: self._set_status("❌ Failed to fetch departments"))
            self.after(0, lambda: self._update_conn_badge(False))

    def _render_departments(self, depts):
        # Clear existing
        for w in self.dept_list_frame.winfo_children():
            w.destroy()

        if not depts:
            ctk.CTkLabel(self.dept_list_frame, text="No departments found.",
                         font=FONT_BODY, text_color=TEXT_SECONDARY).pack(pady=12)
            return

        # Header row
        hdr = ctk.CTkFrame(self.dept_list_frame, fg_color="transparent")
        hdr.pack(fill="x", padx=12, pady=(10, 4))
        ctk.CTkLabel(hdr, text="Department Name", font=FONT_BADGE, text_color=TEXT_PRIMARY, width=200).pack(side="left", padx=4)
        ctk.CTkLabel(hdr, text="First Year Seq", font=FONT_BADGE, text_color=TEXT_PRIMARY, width=120).pack(side="left", padx=4)
        ctk.CTkLabel(hdr, text="Last Year Seq", font=FONT_BADGE, text_color=TEXT_PRIMARY, width=120).pack(side="left", padx=4)
        ctk.CTkLabel(hdr, text="Prefixes (CSV)", font=FONT_BADGE, text_color=TEXT_PRIMARY).pack(side="left", padx=4)

        self.dept_entry_widgets = {} # dept_name -> (first_e, last_e, prefix_e)
        
        # Load existing dept settings
        raw = get_setting("active_settings", "{}")
        try:
            current_settings = json.loads(raw).get("depts", {})
        except:
            current_settings = {}

        for dept in depts:
            dname = dept.get("dept_name", "")
            drow = ctk.CTkFrame(self.dept_list_frame, fg_color=SURFACE, corner_radius=6,
                               border_width=1, border_color=BORDER)
            drow.pack(fill="x", padx=12, pady=1)
            
            ctk.CTkLabel(drow, text=dname, font=FONT_BODY, text_color=TEXT_PRIMARY, width=204, anchor="w").pack(side="left", padx=10)
            
            f_e = ctk.CTkEntry(drow, width=110, font=FONT_SMALL, placeholder_text="e.g. 21CS001")
            f_e.pack(side="left", padx=4, pady=6)
            
            l_e = ctk.CTkEntry(drow, width=110, font=FONT_SMALL, placeholder_text="e.g. 24CS999")
            l_e.pack(side="left", padx=4, pady=6)
            
            p_e = ctk.CTkEntry(drow, width=150, font=FONT_SMALL, placeholder_text="e.g. 21CS,22CS")
            p_e.pack(side="left", padx=4, pady=6)
            
            # Populate if exists
            dset_raw = current_settings.get(dname)
            if dset_raw and isinstance(dset_raw, dict):
                first_seq = dset_raw.get("first_seq")
                last_seq = dset_raw.get("last_seq")
                prefixes_val = dset_raw.get("prefixes")
                if first_seq: f_e.insert(0, str(first_seq))
                if last_seq: l_e.insert(0, str(last_seq))
                if prefixes_val: p_e.insert(0, str(prefixes_val))
            
            self.dept_entry_widgets[dname] = (f_e, l_e, p_e)

        ctk.CTkLabel(self.dept_list_frame, text=f"Total: {len(depts)} departments. Don't forget to Save below!",
                     font=FONT_BADGE, text_color=SUCCESS).pack(pady=(10, 10))

    # ─── status ───────────────────────────────────────────────────────
    def _set_status(self, txt):
        self.status_lbl.configure(text=txt)

    def _update_conn_badge(self, connected):
        """Update the connection badge color and text."""
        if connected:
            self.conn_badge.configure(
                text="  ● Server Connected  ",
                fg_color=SUCCESS, text_color="#ffffff"
            )
        else:
            self.conn_badge.configure(
                text="  ● Disconnected  ",
                fg_color=DANGER, text_color="#ffffff"
            )

    def _check_connection(self):
        """Test DB connection on startup (runs in background thread)."""
        def _worker():
            try:
                conn = mssql_connect()
                if conn: conn.close()
                self.after(0, lambda: self._update_conn_badge(True))
                self.after(0, lambda: self._set_status("Ready"))
            except Exception as e:
                import traceback
                print("\n❌ [ERROR] Startup Connection Check Failed:")
                traceback.print_exc()
                self.after(0, lambda: self._update_conn_badge(False))
                self.after(0, lambda: self._set_status("DB Offline"))

        threading.Thread(target=_worker, daemon=True).start()

    # ═══════════════════════════════════════════════════════════════════
    #  FETCH DATA  (reminder / overdue)
    # ═══════════════════════════════════════════════════════════════════
    def _fetch_reminder(self):
        self._set_status("⏳ Fetching reminder data…")
        threading.Thread(target=self._fetch_reminder_thread, daemon=True).start()

    def _fetch_reminder_thread(self):
        try:
            prefixes = self._get_prefixes()
            rows = fetch_reminder_students(prefixes)
            self.reminder_data = group_by_student(rows)
            self.after(0, self._render_reminder)
            self.after(0, lambda: self._update_conn_badge(True))
        except Exception as e:
            import traceback
            print("\n❌ [ERROR] Fetch Reminder Thread Failed:")
            traceback.print_exc()
            err_msg = str(e)
            self.after(0, lambda: self._set_status(f"❌ DB Error"))
            self.after(0, lambda: self._update_conn_badge(False))
            self.after(0, lambda: messagebox.showerror("Database Error", f"FULL ERROR:\n\n{err_msg}\n\nCheck terminal for full traceback."))

    def _fetch_today(self):
        self._set_status("⏳ Fetching today's due data…")
        threading.Thread(target=self._fetch_today_thread, daemon=True).start()

    def _fetch_today_thread(self):
        try:
            prefixes = self._get_prefixes()
            rows = fetch_today_students(prefixes)
            self.today_data = group_by_student(rows)
            self.after(0, self._render_today)
            self.after(0, lambda: self._update_conn_badge(True))
        except Exception as e:
            import traceback
            print("\n❌ [ERROR] Fetch Today Thread Failed:")
            traceback.print_exc()
            err_msg = str(e)
            self.after(0, lambda: self._set_status(f"❌ DB Error"))
            self.after(0, lambda: self._update_conn_badge(False))
            self.after(0, lambda: messagebox.showerror("Database Error", f"FULL ERROR:\n\n{err_msg}\n\nCheck terminal for full traceback."))

    def _fetch_overdue(self):
        self._set_status("⏳ Fetching overdue data…")
        threading.Thread(target=self._fetch_overdue_thread, daemon=True).start()

    def _fetch_overdue_thread(self):
        try:
            prefixes = self._get_prefixes()
            active_settings = self._get_active_settings()
            rows = fetch_overdue_students(prefixes, active_settings)
            self.overdue_data = group_by_student(rows)
            # collect depts for filter
            depts = sorted(set(d["Department"] for d in self.overdue_data.values() if d["Department"]))
            self.after(0, lambda: self.overdue_dept_filter.configure(values=["All"] + depts))
            self.after(0, self._render_overdue)
            self.after(0, lambda: self._update_conn_badge(True))
        except Exception as e:
            import traceback
            print("\n❌ [ERROR] Fetch Overdue Thread Failed:")
            traceback.print_exc()
            err_msg = str(e)
            self.after(0, lambda: self._set_status(f"❌ DB Error"))
            self.after(0, lambda: self._update_conn_badge(False))
            self.after(0, lambda: messagebox.showerror("Database Error", f"FULL ERROR:\n\n{err_msg}\n\nCheck terminal for full traceback."))

    # ═══════════════════════════════════════════════════════════════════
    #  SEARCH AND EXPORT
    # ═══════════════════════════════════════════════════════════════════
    def _trigger_search(self):
        """Called when the search entry text changes or tab switches."""
        if not hasattr(self, 'current_tab'):
            return
        if self.current_tab == "reminder":
            self._render_reminder()
        elif self.current_tab == "today":
            self._render_today()
        elif self.current_tab == "overdue":
            self._sort_overdue_display()

    def _export_to_excel(self, mode):
        """Export currently viewed and filtered data to an Excel file with styling."""
        if not hasattr(self, 'current_tab'):
            return
            
        if self.current_tab == "reminder":
            data = self.reminder_data
        elif self.current_tab == "today":
            data = self.today_data
        elif self.current_tab == "overdue":
            data = self.overdue_data
        else:
            messagebox.showinfo("Export", "Please navigate to Reminder, Today's Due, or Overdue to export.")
            return

        q = self.search_var.get().lower().strip()
        # Filter by search string
        filtered_data = {roll: info for roll, info in data.items() 
                        if (q in str(roll).lower() or q in info["Name"].lower())
                        and info.get("MemberType") != "Teacher"} # EXCLUDE TEACHERS

        # Apply department filter if on overdue tab
        if self.current_tab == "overdue":
            dept_filter = self.overdue_dept_filter.get()
            if dept_filter != "All":
                filtered_data = {k: v for k, v in filtered_data.items() if v["Department"] == dept_filter}

        if not filtered_data:
            messagebox.showwarning("No Data", "No matching student data to export (Teachers are excluded).")
            return

        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All Files", "*.*")],
            title="Export to Excel"
        )
        if not filepath:
            return

        try:
            # Prepare rows
            rows = []
            serial_no = 1
            for roll, info in sorted(filtered_data.items(), key=lambda x: (x[1]["Department"], x[1]["Name"])):
                dept = info.get("Department", "No Department")
                
                # Mode-based row creation
                if mode == "students_only":
                    # Image 1: S.NO, RGE.NO (RGE.NO matches typo in user image), NAME, DEPART
                    rows.append({
                        "S.NO": serial_no,
                        "RGE.NO": roll,
                        "NAME": info["Name"],
                        "DEPART": dept
                    })
                    serial_no += 1
                elif mode == "detailed":
                    # Image 2: REG .NO, NAME, ACC .NO, BOOK NAME, AUTHOR NAM, ISSUE DATE, RET DATE
                    for b in info["books"]:
                        # Formatting dates
                        issue_dt = b.get("IssueDate")
                        if isinstance(issue_dt, datetime.datetime): issue_dt = issue_dt.strftime("%d-%m-%Y")
                        elif isinstance(issue_dt, datetime.date): issue_dt = issue_dt.strftime("%d-%m-%Y")
                        
                        due_dt = b.get("DueDate")
                        if isinstance(due_dt, datetime.datetime): due_dt = due_dt.strftime("%d-%m-%Y")
                        elif isinstance(due_dt, datetime.date): due_dt = due_dt.strftime("%d-%m-%Y")

                        rows.append({
                            "REG .NO": roll,
                            "NAME": info["Name"],
                            "ACC .NO": b.get("AccNo", ""),
                            "BOOK NAME": b.get("BookTitle", ""),
                            "AUTHOR NAM": b.get("Author", ""),
                            "ISSUE DATE": issue_dt or "",
                            "RET DATE": due_dt or ""
                        })
                else: # all_cols (Standard Summary + Book/Author)
                    total_fine = sum(b.get("FineAmount", 0) for b in info["books"]) if self.current_tab == "overdue" else 0
                    h_count = len(self.history_map.get(roll, []))
                    
                    for b in info["books"]:
                        rows.append({
                            "Department": dept,
                            "Reg No": roll,
                            "Name": info["Name"],
                            "Book Title": b.get("BookTitle", ""),
                            "Author Name": b.get("Author", ""),
                            "Email": info["Email"],
                            "Member Type": info.get("MemberType", ""),
                            "Books Count": len(info["books"]),
                            "Total Fine": total_fine,
                            "History Count": h_count
                        })

            df = pd.DataFrame(rows)
            
            # Export using pandas + openpyxl for styling
            # Using a more robust method for Windows 7 / WPS compatibility
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 🛡️ FIX: Ensure at least one sheet exists and is active to avoid "visonal sheet" errors
                workbook = writer.book
                sheet_name = "LibraryExport"
                if sheet_name in workbook.sheetnames:
                    std = workbook[sheet_name]
                else:
                    std = workbook.create_sheet(sheet_name)
                workbook.active = std

                start_row = 0
                # Group by department to add headers
                for dept, dept_df in df.groupby("Department", sort=False):
                    # ── Write Department Header ──
                    title_row_idx = start_row + 1
                    
                    # Merge across all columns
                    std.merge_cells(start_row=title_row_idx, start_column=1, end_row=title_row_idx, end_column=len(df.columns))
                    
                    title_cell = std.cell(row=title_row_idx, column=1)
                    title_cell.value = f"DEPARTMENT: {dept}"
                    
                    # Styling: Text Red, Background Yellow, Centered
                    title_cell.font = Font(color="FF0000", bold=True, size=14)
                    title_cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
                    title_cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # ── Write Data for this department ──
                    # to_excel startrow is 0-indexed in pandas writer
                    dept_df.to_excel(writer, index=False, startrow=title_row_idx, sheet_name=sheet_name)
                    
                    # Increment start_row for next group: 
                    # 1 row for Dept Title + 1 row for Table Header + Data Rows + 1 Spacer
                    start_row += 1 + 1 + len(dept_df) + 1

                # 🛡️ FIX: Remove the default empty "Sheet" if created by openpyxl
                if "Sheet" in workbook.sheetnames and len(workbook.sheetnames) > 1:
                    del workbook["Sheet"]
            
            self._set_status(f"✅ Exported successfully to {os.path.basename(filepath)}")
            messagebox.showinfo("Export Success", f"Successfully exported {len(filtered_data)} records to Excel.")
        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Export Failed", f"Failed to export Excel: {str(e)}")


    # ═══════════════════════════════════════════════════════════════════
    #  RENDER CARDS
    # ═══════════════════════════════════════════════════════════════════
    def _clear_scroll(self, scroll_frame, widget_list):
        for w in widget_list:
            w.destroy()
        widget_list.clear()

    def _render_reminder(self):
        self._clear_scroll(self.reminder_scroll, self.reminder_widgets)
        self.reminder_checks.clear()
        if hasattr(self, 'reminder_empty'):
            self.reminder_empty.destroy()

        data = self.reminder_data
        if not data:
            lbl = ctk.CTkLabel(self.reminder_scroll, text="✅ No students have books due tomorrow.",
                               font=FONT_BODY, text_color=SUCCESS)
            lbl.pack(pady=40)
            self.reminder_widgets.append(lbl)
            self.reminder_count_lbl.configure(text="0 students")
            self._set_status("✅ Reminder: 0 students")
            return

        q = self.search_var.get().lower().strip()
        filtered_data = {roll: info for roll, info in data.items() if q in str(roll).lower() or q in info["Name"].lower()}

        for roll, info in sorted(filtered_data.items(), key=lambda x: x[1]["Name"]):
            self._render_student_card(self.reminder_scroll, info, "reminder")

        self.reminder_count_lbl.configure(text=f"{len(filtered_data)} student(s)")
        self._set_status(f"✅ Loaded {len(filtered_data)} reminder student(s)")

    def _render_today(self):
        self._clear_scroll(self.today_scroll, self.today_widgets)
        self.today_checks.clear()
        if hasattr(self, 'today_empty'):
            self.today_empty.destroy()

        data = self.today_data
        if not data:
            lbl = ctk.CTkLabel(self.today_scroll, text="✅ No students have books due today.",
                               font=FONT_BODY, text_color=SUCCESS)
            lbl.pack(pady=40)
            self.today_widgets.append(lbl)
            self.today_count_lbl.configure(text="0 students")
            self._set_status("✅ Today: 0 students")
            return

        q = self.search_var.get().lower().strip()
        filtered_data = {roll: info for roll, info in data.items() if q in str(roll).lower() or q in info["Name"].lower()}

        for roll, info in sorted(filtered_data.items(), key=lambda x: x[1]["Name"]):
            self._render_student_card(self.today_scroll, info, "today")

        self.today_count_lbl.configure(text=f"{len(filtered_data)} student(s)")
        self._set_status(f"✅ Loaded {len(filtered_data)} today due student(s)")

    def _render_overdue(self):
        self._sort_overdue_display()

    def _on_overdue_tab_change(self):
        # Update status or count when tab changes
        pass

    def _sort_overdue_display(self, _event=None):
        # Clear all lists
        self._clear_scroll(self.active_scroll, self.overdue_widgets)
        self._clear_scroll(self.passout_scroll, self.overdue_widgets)
        self._clear_scroll(self.teacher_scroll, self.overdue_widgets)
        
        self.overdue_checks.clear()
        
        data = dict(self.overdue_data)  # copy
        if not data:
            self.overdue_count_lbl.configure(text="0 students")
            self._set_status("✅ Overdue: 0 students")
            return

        # apply search filter
        q = self.search_var.get().lower().strip()
        if q:
            data = {roll: info for roll, info in data.items() if q in str(roll).lower() or q in info["Name"].lower()}

        # apply dept filter
        dept_filter = self.overdue_dept_filter.get()
        if dept_filter != "All":
            data = {k: v for k, v in data.items() if v["Department"] == dept_filter}

        # apply day range
        try:
            min_d = int(self.overdue_min_days.get()) if self.overdue_min_days.get().strip() else 0
        except ValueError:
            min_d = 0
        try:
            max_d = int(self.overdue_max_days.get()) if self.overdue_max_days.get().strip() else 999999
        except ValueError:
            max_d = 999999

        def max_overdue(info):
            return max((b.get("OverdueDays", 0) for b in info["books"]), default=0)

        data = {k: v for k, v in data.items() if min_d <= max_overdue(v) <= max_d}

        # sort
        sort_key = self.overdue_sort.get()
        if sort_key.startswith("Name"):
            items = sorted(data.items(), key=lambda x: x[1]["Name"])
        elif sort_key.startswith("Days"):
            items = sorted(data.items(), key=lambda x: max_overdue(x[1]), reverse=True)
        elif sort_key.startswith("Fine"):
            items = sorted(data.items(), key=lambda x: sum(b.get("FineAmount", 0) for b in x[1]["books"]), reverse=True)
        elif sort_key.startswith("Year"):
            def get_year(roll):
                y = "".join([c for c in str(roll) if c.isdigit()])[:2]  # type: ignore
                return int(y) if y else 0
            items = sorted(data.items(), key=lambda x: get_year(x[1]["RollNumber"]), reverse=True)
        elif sort_key.startswith("Department"):
            items = sorted(data.items(), key=lambda x: x[1]["Department"])
        elif sort_key.startswith("Book"):
            items = sorted(data.items(), key=lambda x: x[1]["books"][0].get("BookTitle", "") if x[1]["books"] else "")
        else:
            items = list(data.items())

        counts = {"Active Student": 0, "Passout Student": 0, "Teacher": 0}

        for roll, info in items:
            mtype = info.get("MemberType", "Active Student")
            target_scroll = self.active_scroll
            if mtype == "Passout Student":
                target_scroll = self.passout_scroll
            elif mtype == "Teacher":
                target_scroll = self.teacher_scroll
            
            self._render_student_card(target_scroll, info, "overdue")
            counts[mtype] += 1

        # Update status label with detailed breakdown instead of trying to rename tabs (which CTkTabview doesn't support well)
        detail_txt = f"Total: {len(items)} | Active: {counts['Active Student']} | Passout: {counts['Passout Student']} | Teacher: {counts['Teacher']}"
        self.overdue_count_lbl.configure(text=detail_txt)
        self._set_status(f"✅ Overdue: {len(items)} students loaded")

    def _render_student_card(self, parent, info, tab_type):
        """Render a single student card with checkbox."""
        roll = info["RollNumber"]
        already_sent = was_sent_today(roll, tab_type)

        # For "Today" tab, also check if reminded YESTERDAY
        reminded_yesterday = False
        if tab_type == "today":
            yesterday = datetime.date.today() - datetime.timedelta(days=1)
            reminded_yesterday = was_sent_on(roll, "reminder", yesterday)

        card = ctk.CTkFrame(parent, fg_color=SURFACE, corner_radius=10,
                            border_width=1, border_color="#e2e8f0" if not (already_sent or reminded_yesterday) else "#bbf7d0")
        card.pack(fill="x", padx=4, pady=3)

        if tab_type == "reminder":
            self.reminder_widgets.append(card)
        elif tab_type == "today":
            self.today_widgets.append(card)
        else:
            self.overdue_widgets.append(card)

        inner = ctk.CTkFrame(card, fg_color="transparent")
        inner.pack(fill="x", padx=12, pady=10)

        # checkbox
        var = ctk.BooleanVar(value=False)
        chk = ctk.CTkCheckBox(inner, text="", variable=var, width=24,
                               checkbox_width=22, checkbox_height=22,
                               fg_color=PRIMARY, hover_color=PRIMARY_HOVER)
        chk.pack(side="left", padx=(0, 10))

        if tab_type == "reminder":
            self.reminder_checks[roll] = var
        elif tab_type == "today":
            self.today_checks[roll] = var
        else:
            self.overdue_checks[roll] = var

        # info
        info_frame = ctk.CTkFrame(inner, fg_color="transparent")
        info_frame.pack(side="left", fill="x", expand=True)

        top_row = ctk.CTkFrame(info_frame, fg_color="transparent")
        top_row.pack(fill="x")

        ctk.CTkLabel(top_row, text=info["Name"], font=("Segoe UI", 14, "bold"),
                     text_color=TEXT_PRIMARY).pack(side="left")
        ctk.CTkLabel(top_row, text=f"  {roll}", font=FONT_SMALL,
                     text_color=TEXT_SECONDARY).pack(side="left", padx=(6, 0))
        ctk.CTkLabel(top_row, text=f"  |  {info['Department']}", font=FONT_SMALL,
                     text_color=TEXT_SECONDARY).pack(side="left", padx=(4, 0))

        if already_sent:
            ctk.CTkLabel(top_row, text="  ✅ Sent today", font=FONT_BADGE,
                         text_color=SUCCESS).pack(side="left", padx=(8, 0))
        elif reminded_yesterday:
            ctk.CTkLabel(top_row, text="  ✅ Reminded yesterday", font=FONT_BADGE,
                         text_color="#10b981").pack(side="left", padx=(8, 0))
        elif tab_type == "today":
            # Highlight if missed reminder yesterday
            ctk.CTkLabel(top_row, text="  ⚠️ Missed reminder", font=FONT_BADGE,
                         text_color=WARNING).pack(side="left", padx=(8, 0))
                         
        # Tag label showing history count
        history_dates = self.history_map.get(roll, [])
        if history_dates:
            history_count = len(history_dates)
            last_sent = history_dates[0]  # First element is most recent because of ORDER BY DESC
            hist_lbl = ctk.CTkLabel(top_row, text=f"  🔄 Sent: {history_count}", font=FONT_BADGE,
                         text_color=PRIMARY)
            hist_lbl.pack(side="left", padx=(8, 0))
            # Adding tooltip/details for sent dates when clicked using CTkButton or just displaying last sent
            ctk.CTkLabel(top_row, text=f" (Last: {last_sent[:10]})", font=FONT_SMALL,
                         text_color=TEXT_SECONDARY).pack(side="left")

        # email
        ctk.CTkLabel(info_frame, text=f"📧 {info['Email']}", font=FONT_SMALL,
                     text_color=TEXT_SECONDARY).pack(anchor="w", pady=(2, 0))

        # books
        books = info["books"]
        for b in books:
            due = b.get("DueDate", "")
            if isinstance(due, datetime.datetime):
                due = due.strftime("%d %b %Y")
            txt = f"📖 {b.get('BookTitle', 'N/A')}  |  Acc: {b.get('AccNo', '')}  |  Due: {due}"
            if tab_type == "overdue":
                txt += f"  |  Overdue: {b.get('OverdueDays', 0)}d  |  Fine: ₹{b.get('FineAmount', 0)}"
            ctk.CTkLabel(info_frame, text=txt, font=FONT_SMALL,
                         text_color="#475569").pack(anchor="w", pady=1)

        # total fine for overdue
        if tab_type == "overdue":
            total_fine = sum(b.get("FineAmount", 0) for b in books)
            if total_fine > 0:
                ctk.CTkLabel(info_frame, text=f"💰 Total Fine: ₹{total_fine:.0f}",
                             font=("Segoe UI", 12, "bold"), text_color=DANGER).pack(anchor="w", pady=(2, 0))

    # ═══════════════════════════════════════════════════════════════════
    #  SELECT / DESELECT
    # ═══════════════════════════════════════════════════════════════════
    def _select_all(self, tab_type):
        if tab_type == "reminder": checks = self.reminder_checks
        elif tab_type == "today": checks = self.today_checks
        else:
            # Overdue: only select members in the currently VISIBLE sub-tab
            current_sub = self.overdue_tabs.get()
            mtype_map = {"Active Students": "Active Student", "Passout Students": "Passout Student", "Teachers": "Teacher"}
            target_type = mtype_map.get(current_sub)
            
            for roll, var in self.overdue_checks.items():
                if self.overdue_data[roll]["MemberType"] == target_type:
                    var.set(True)
            return

        for v in checks.values():
            v.set(True)

    def _deselect_all(self, tab_type):
        if tab_type == "reminder": checks = self.reminder_checks
        elif tab_type == "today": checks = self.today_checks
        else:
            # Overdue: only deselect members in the currently VISIBLE sub-tab
            current_sub = self.overdue_tabs.get()
            mtype_map = {"Active Students": "Active Student", "Passout Students": "Passout Student", "Teachers": "Teacher"}
            target_type = mtype_map.get(current_sub)
            
            for roll, var in self.overdue_checks.items():
                if self.overdue_data[roll]["MemberType"] == target_type:
                    var.set(False)
            return

        for v in checks.values():
            v.set(False)

    # ═══════════════════════════════════════════════════════════════════
    #  SEND EMAILS
    # ═══════════════════════════════════════════════════════════════════
    def _send_selected(self, tab_type):
        if tab_type == "reminder":
            checks, data = self.reminder_checks, self.reminder_data
        elif tab_type == "today":
            checks, data = self.today_checks, self.today_data
        else:
            checks, data = self.overdue_checks, self.overdue_data

        selected = [roll for roll, var in checks.items() if var.get()]
        if not selected:
            messagebox.showwarning("No Selection", "Please select at least one student.")
            return
        self._confirm_and_send(tab_type, data, selected)

    def _send_all(self, tab_type):
        if tab_type == "reminder": data = self.reminder_data
        elif tab_type == "today": data = self.today_data
        else: data = self.overdue_data

        if not data:
            messagebox.showwarning("No Data", "No students loaded. Fetch data first.")
            return
        selected = list(data.keys())
        self._confirm_and_send(tab_type, data, selected)

    def _confirm_and_send(self, tab_type, data, roll_list):
        # skip already sent today
        not_sent = [r for r in roll_list if not was_sent_today(r, tab_type)]
        already = len(roll_list) - len(not_sent)

        msg = f"Send {tab_type.title()} email to {len(not_sent)} student(s)?"
        if already:
            msg += f"\n({already} already sent today and will be skipped.)"
        if not not_sent:
            messagebox.showinfo("Already Sent", "All selected students were already emailed today.")
            return
        if not messagebox.askyesno("Confirm Send", msg):
            return

        self._set_status(f"📤 Sending {len(not_sent)} emails…")
        threading.Thread(target=self._send_thread, args=(tab_type, data, not_sent), daemon=True).start()

    def _send_thread(self, tab_type, data, roll_list):
        ok: int = 0
        fail: int = 0
        for roll in roll_list:
            info = data.get(roll)
            if not info or not info["Email"]:
                continue
            try:
                if tab_type == "reminder":
                    subject = f"📅 Book Due Tomorrow — {info['books'][0]['BookTitle']}"
                    html = build_reminder_html(info["Name"], info["books"])
                elif tab_type == "today":
                    subject = f"⏰ Book Due TODAY — {info['books'][0]['BookTitle']}"
                    html = build_reminder_html(info["Name"], info["books"])
                    # Use a slightly different title in HTML for "Today"
                    html = html.replace("Due Date Reminder", "Due TODAY Notice")
                    html = html.replace("due <strong>tomorrow</strong>", "due <strong>TODAY</strong>")
                else:
                    total_fine = sum(b.get("FineAmount", 0) for b in info["books"])
                    subject = f"⚠️ Overdue Alert — Fine: ₹{total_fine:.0f}"
                    html = build_overdue_html(info["Name"], info["books"], f"{total_fine:.0f}")

                send_email(info["Email"], subject, html)
                # Store as 'today_reminder' to distinguish if needed
                mtype_to_save = "today_reminder" if tab_type == "today" else tab_type
                add_sent_record(roll, info["Name"], info["Email"], mtype_to_save, subject,
                                json.dumps([b.get("BookTitle", "") for b in info["books"]]))
                ok = ok + 1  # type: ignore
            except Exception as e:
                fail = fail + 1  # type: ignore
                print(f"❌ Email failed for {roll}: {e}")

        self.after(0, lambda: self._set_status(f"✅ Sent: {ok}  |  Failed: {fail}"))
        self.after(0, lambda: messagebox.showinfo("Done", f"Sent: {ok}\nFailed: {fail}"))
        
        # update history map and UI
        self.history_map = get_all_sent_history_counts()
        
        # refresh cards to show sent status
        if tab_type == "reminder":
            self.after(100, self._render_reminder)
        elif tab_type == "today":
            self.after(100, self._render_today)
        else:
            self.after(100, self._render_overdue)

    # ═══════════════════════════════════════════════════════════════════
    #  HISTORY
    # ═══════════════════════════════════════════════════════════════════
    def _load_history(self):
        self._clear_scroll(self.history_scroll, self.history_widgets)
        self.history_checks.clear()
        rows = get_history(300)
        self.history_count_lbl.configure(text=f"{len(rows)} records")

        if not rows:
            lbl = ctk.CTkLabel(self.history_scroll, text="No history yet.",
                               font=FONT_BODY, text_color=TEXT_SECONDARY)
            lbl.pack(pady=40)
            self.history_widgets.append(lbl)
            return

        for row in rows:
            hid, roll, name, email, mtype, subj, sent_at, details = row
            card = ctk.CTkFrame(self.history_scroll, fg_color=SURFACE, corner_radius=8,
                                border_width=1, border_color=BORDER)
            card.pack(fill="x", padx=4, pady=2)
            self.history_widgets.append(card)

            inner = ctk.CTkFrame(card, fg_color="transparent")
            inner.pack(fill="x", padx=10, pady=8)

            var = ctk.BooleanVar(value=False)
            chk = ctk.CTkCheckBox(inner, text="", variable=var, width=24,
                                   checkbox_width=20, checkbox_height=20,
                                   fg_color=PRIMARY, hover_color=PRIMARY_HOVER)
            chk.pack(side="left", padx=(0, 8))
            self.history_checks[hid] = var

            info = ctk.CTkFrame(inner, fg_color="transparent")
            info.pack(side="left", fill="x", expand=True)

            badge_color = WARNING if mtype == "reminder" else DANGER
            badge_text = "📅 Reminder" if mtype == "reminder" else "⚠️ Overdue"

            top = ctk.CTkFrame(info, fg_color="transparent")
            top.pack(fill="x")
            ctk.CTkLabel(top, text=badge_text, font=FONT_BADGE, text_color=badge_color).pack(side="left")
            ctk.CTkLabel(top, text=f"  {name}  ({roll})", font=FONT_BODY,
                         text_color=TEXT_PRIMARY).pack(side="left", padx=(8, 0))
            ctk.CTkLabel(top, text=f"  📧 {email}", font=FONT_SMALL,
                         text_color=TEXT_SECONDARY).pack(side="left", padx=(8, 0))
            ctk.CTkLabel(top, text=sent_at, font=FONT_SMALL,
                         text_color=TEXT_SECONDARY).pack(side="right")

            ctk.CTkLabel(info, text=subj, font=FONT_SMALL,
                         text_color="#475569").pack(anchor="w", pady=(2, 0))

    def _select_all_history(self):
        for v in self.history_checks.values():
            v.set(True)

    def _delete_selected_history(self):
        selected = [hid for hid, var in self.history_checks.items() if var.get()]
        if not selected:
            messagebox.showwarning("No Selection", "Select history records to delete.")
            return
        if messagebox.askyesno("Confirm Delete", f"Delete {len(selected)} record(s)?"):
            delete_history_ids(selected)
            self._load_history()
            self._set_status(f"🗑️ Deleted {len(selected)} record(s)")


# ══════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    app = App()
    app.mainloop()
